"""
v2 的执行器、校验器、修复器一站式实现。

本文件是 v2 的唯一核心基础设施，没有任何对外部 legacy 模块的依赖。

包含：
- _patch_reportlab_cjk_support: reportlab 中文字体运行时补丁
- _execute_python_code: 双引擎（E2B / 本地子进程），含错误分类与产物收集
- executor_node / validator_node / fixer_node: 三个 LangGraph 风格节点
- fixer_agent: 让 LLM 根据诊断修复代码

LangGraph 风格的 state dict 与旧版完全兼容，方便沿用旧测试。
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import uuid
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from pathlib import Path
from typing import Any, Optional

from e2b_code_interpreter import Sandbox
from langchain_core.messages import AIMessage, HumanMessage, SystemMessage

from backend.agent.v2.llm import _create_llm
from backend.app_paths import (
    ARTIFACTS_DIR,
    BACKEND_DIR,
    DATA_DIR,
    DUCKDB_PATH,
    artifact_relpath,
    ensure_directories,
)


__all__ = [
    "MAX_GRAPH_RETRIES",
    "LOCAL_EXEC_TIMEOUT_SECONDS",
    "LOCAL_ONLY_MARKER",
    "CJK_FONT_HELPER_MARKER",
    "_patch_reportlab_cjk_support",
    "_classify_error",
    "_strip_artifact_markers",
    "_validate_execution",
    "_execute_python_code",
    "executor_node",
    "validator_node",
    "fixer_node",
    "fixer_agent",
]


# ── 常量 ───────────────────────────────────────────────────────
MAX_GRAPH_RETRIES = 3
LOCAL_EXEC_TIMEOUT_SECONDS = 60
LOCAL_RUNNER_PATH = Path(__file__).resolve().parent.parent.parent.parent / "local_exec_runner.py"
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
SEARCH_SERVICE_PATH = BACKEND_DIR / "search_service.py"
LOCAL_ONLY_MARKER = "# EXECUTION_BACKEND:LOCAL_ONLY"
CJK_FONT_HELPER_MARKER = "# CJK_FONT_RUNTIME_PATCH_V1"


def _get_task_goal(state: dict) -> str:
    task_goal = state.get("task_goal")
    if task_goal:
        return task_goal
    human_messages = [m.content for m in state.get("messages", []) if isinstance(m, HumanMessage)]
    return human_messages[-1] if human_messages else ""


def _make_run_dir(thread_id: str, attempt_index: int) -> Path:
    run_dir = DATA_DIR / "runs" / thread_id / f"attempt_{attempt_index}_{uuid.uuid4().hex[:8]}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def _copy_duckdb_to_workdir(work_dir: Path) -> None:
    if DUCKDB_PATH.exists():
        shutil.copy2(DUCKDB_PATH, work_dir / "data.duckdb")


def _sync_duckdb_back(work_dir: Path) -> None:
    run_db = work_dir / "data.duckdb"
    if run_db.exists():
        shutil.copy2(run_db, DUCKDB_PATH)


def _classify_error(error_type: str, diagnostic_text: str) -> str:
    text = f"{error_type}\n{diagnostic_text}".lower()
    if "syntaxerror" in text or "indentationerror" in text:
        return "syntax_error"
    if "timeout" in text:
        return "timeout"
    if "modulenotfounderror" in text or "importerror" in text:
        return "missing_import"
    if "catalog" in text and "does not exist" in text:
        return "missing_table"
    if "binder" in text or "column" in text and "does not exist" in text:
        return "missing_column"
    if "permission" in text or "access is denied" in text:
        return "filesystem_permission"
    if "plotly" in text or "kaleido" in text:
        return "plot_render_error"
    return "runtime_error"


def _artifact_type_from_name(name: str) -> tuple[Optional[str], Optional[str]]:
    suffix = Path(name).suffix.lower()
    if suffix == ".html":
        return "chart_html", "chart"
    if suffix == ".png":
        return "chart_png", "chart"
    if suffix in {".xlsx", ".xls"}:
        return "cleaned_data", "cleaned"
    if suffix == ".docx":
        return "report_docx", "report"
    if suffix == ".json":
        lower_name = Path(name).name.lower()
        if "validation" in lower_name:
            return "validation_report", "report"
    if suffix == ".pdf":
        return "report_pdf", "report"
    if suffix == ".pptx":
        return "report_pptx", "report"
    return None, None


def _register_artifact(artifacts: list[dict], item_type: str, rel_path: str) -> None:
    if any(existing["path"] == rel_path for existing in artifacts):
        return
    artifacts.append({
        "type": item_type,
        "path": rel_path,
        "name": os.path.basename(rel_path),
    })


def _move_generated_file(source: Path, prefix: str) -> str:
    target = ARTIFACTS_DIR / f"{prefix}_{uuid.uuid4().hex[:8]}{source.suffix}"
    shutil.move(str(source), target)
    return artifact_relpath(target)


def _collect_local_artifacts(work_dir: Path, stdout_text: str, generated_files: list[str]) -> dict:
    chart_path = None
    chart_png_path = None
    cleaned_file_path = None
    report_path = None
    pdf_report_path = None
    pptx_report_path = None
    validation_report_path = None
    artifacts: list[dict] = []
    normalized_output = stdout_text
    seen_relpaths: set[str] = set()

    def move_and_record(rel_path: str, forced_type: Optional[str] = None) -> Optional[str]:
        nonlocal chart_path, chart_png_path, cleaned_file_path, report_path, pdf_report_path, pptx_report_path, validation_report_path
        rel_path = rel_path.strip().replace("\\", "/")
        if not rel_path or rel_path in seen_relpaths:
            return None
        source = (work_dir / rel_path).resolve()
        try:
            source.relative_to(work_dir.resolve())
        except ValueError:
            return None
        if not source.exists() or not source.is_file():
            return None

        item_type, prefix = _artifact_type_from_name(source.name)
        if forced_type:
            item_type = forced_type
        if not item_type or not prefix:
            return None

        rel = _move_generated_file(source, prefix)
        seen_relpaths.add(rel_path)
        _register_artifact(artifacts, item_type, rel)
        if item_type == "chart_html" and not chart_path:
            chart_path = rel
        elif item_type == "chart_png" and not chart_png_path:
            chart_png_path = rel
        elif item_type == "cleaned_data" and not cleaned_file_path:
            cleaned_file_path = rel
        elif item_type == "report_docx" and not report_path:
            report_path = rel
        elif item_type == "report_pdf" and not pdf_report_path:
            pdf_report_path = rel
        elif item_type == "report_pptx" and not pptx_report_path:
            pptx_report_path = rel
        elif item_type == "validation_report" and not validation_report_path:
            validation_report_path = rel
        return rel

    marker_map = {
        "CHART_GENERATED:": "chart_html",
        "CHART_PNG_GENERATED:": "chart_png",
        "CLEANED_DATA_GENERATED:": "cleaned_data",
    }
    for marker, forced_type in marker_map.items():
        for match in re.findall(rf"{re.escape(marker)}([^\r\n]+)", stdout_text):
            moved = move_and_record(match, forced_type)
            if moved:
                normalized_output = normalized_output.replace(f"{marker}{match}", f"{marker}{moved}")

    for match in re.findall(r"REPORT_GENERATED:([^\r\n]+)", stdout_text):
        moved = move_and_record(match)
        if moved:
            normalized_output = normalized_output.replace(f"REPORT_GENERATED:{match}", f"REPORT_GENERATED:{moved}")

    for rel_path in generated_files:
        move_and_record(rel_path)

    return {
        "chart_path": chart_path,
        "chart_png_path": chart_png_path,
        "cleaned_file_path": cleaned_file_path,
        "report_path": report_path,
        "pdf_report_path": pdf_report_path,
        "pptx_report_path": pptx_report_path,
        "validation_report_path": validation_report_path,
        "artifacts": artifacts,
        "normalized_output": normalized_output,
    }


def _collect_e2b_artifacts(sandbox: Sandbox, output_text: str) -> dict:
    chart_path = None
    chart_png_path = None
    cleaned_file_path = None
    report_path = None
    pdf_report_path = None
    pptx_report_path = None
    validation_report_path = None
    artifacts: list[dict] = []
    normalized_output = output_text

    marker_handlers = {
        "CHART_GENERATED:": ("chart_html", "chart", "text"),
        "CHART_PNG_GENERATED:": ("chart_png", "chart", "bytes"),
        "CLEANED_DATA_GENERATED:": ("cleaned_data", "cleaned", "bytes"),
    }

    for marker, (item_type, prefix, mode) in marker_handlers.items():
        for sandbox_path in re.findall(rf"{re.escape(marker)}([^\r\n]+)", output_text):
            sandbox_path = sandbox_path.strip()
            try:
                target_suffix = Path(sandbox_path).suffix or {
                    "chart_html": ".html",
                    "chart_png": ".png",
                    "cleaned_data": ".xlsx",
                }[item_type]
                local_path = ARTIFACTS_DIR / f"{prefix}_{uuid.uuid4().hex[:8]}{target_suffix}"
                if mode == "text":
                    content = sandbox.files.read(sandbox_path)
                    local_path.write_text(content, encoding="utf-8")
                else:
                    content = sandbox.files.read_bytes(sandbox_path)
                    local_path.write_bytes(content)
                rel = artifact_relpath(local_path)
                normalized_output = normalized_output.replace(f"{marker}{sandbox_path}", f"{marker}{rel}")
                _register_artifact(artifacts, item_type, rel)
                if item_type == "chart_html" and not chart_path:
                    chart_path = rel
                elif item_type == "chart_png" and not chart_png_path:
                    chart_png_path = rel
                elif item_type == "cleaned_data" and not cleaned_file_path:
                    cleaned_file_path = rel
            except Exception:
                continue

    for sandbox_path in re.findall(r"REPORT_GENERATED:([^\r\n]+)", output_text):
        sandbox_path = sandbox_path.strip()
        try:
            suffix = Path(sandbox_path).suffix.lower()
            if suffix not in {".pdf", ".docx", ".pptx"}:
                continue
            prefix = "report"
            local_path = ARTIFACTS_DIR / f"{prefix}_{uuid.uuid4().hex[:8]}{suffix}"
            local_path.write_bytes(sandbox.files.read_bytes(sandbox_path))
            rel = artifact_relpath(local_path)
            normalized_output = normalized_output.replace(f"REPORT_GENERATED:{sandbox_path}", f"REPORT_GENERATED:{rel}")
            item_type_map = {".pdf": "report_pdf", ".docx": "report_docx", ".pptx": "report_pptx"}
            item_type = item_type_map.get(suffix, "report_docx")
            _register_artifact(artifacts, item_type, rel)
            if item_type == "report_pdf" and not pdf_report_path:
                pdf_report_path = rel
            elif item_type == "report_docx" and not report_path:
                report_path = rel
            elif item_type == "report_pptx" and not pptx_report_path:
                pptx_report_path = rel
        except Exception:
            continue

    return {
        "chart_path": chart_path,
        "chart_png_path": chart_png_path,
        "cleaned_file_path": cleaned_file_path,
        "report_path": report_path,
        "pdf_report_path": pdf_report_path,
        "pptx_report_path": pptx_report_path,
        "validation_report_path": validation_report_path,
        "artifacts": artifacts,
        "normalized_output": normalized_output,
    }


def _preflight_check(code: str) -> Optional[str]:
    forbidden_patterns = [
        r"!\s*pip\s+install",
        r"pip\s+install",
        r"subprocess\.(run|call|Popen)\s*\(",
        r"os\.system\s*\(",
        # 文件系统探索
        r"\bos\.listdir\b",
        r"\bos\.walk\b",
        r"\bos\.scandir\b",
        r"\bglob\.glob\b",
        r"\.glob\s*\([\"\']",
        r"\.rglob\s*\(",
        # DuckDB 系统表探索
        r"information_schema\s*\.\s*tables",
        r"information_schema\s*\.\s*columns",
        r"\bduckdb_tables\s*\(",
        r"\bduckdb_columns\s*\(",
        r"\bduckdb_databases\s*\(",
        r"\bpg_catalog\b",
        r"\bsqlite_master\b",
    ]
    for pattern in forbidden_patterns:
        if re.search(pattern, code, flags=re.IGNORECASE):
            return (
                "生成的代码包含被禁止的操作（系统命令执行、文件系统浏览或数据库系统表查询）。"
                "请直接使用当前环境已有依赖完成任务，不要调用 pip、subprocess、os.system，"
                "不要浏览文件系统（os.listdir/os.walk/glob），"
                "不要直接查询数据库系统表。使用 list_tables / query_duckdb 工具完成数据访问。"
            )
    return None


def _preflight_researcher_check(code: str) -> Optional[str]:
    """researcher agent 的搜索已在 agent 层预执行，代码只需处理预获取数据。"""
    if "from search_service" in code or "import search_service" in code:
        return (
            "你的代码不应导入 search_service 模块。"
            "搜索已由系统预执行，结果已固化为 _SEARCH_RESULTS 和 _CRAWLED_RESULTS 字面量。"
            "请删除 `from search_service import ...` 语句，直接使用提示词中提供的预定义数据。"
        )
    forbidden_in_researcher = [
        (r"urllib\.request\.urlopen\s*\(", "urllib.request.urlopen"),
        (r"urllib\.request\.Request\s*\(", "urllib.request.Request"),
        (r"HTMLParser\(", "HTMLParser"),
        (r"BeautifulSoup\(", "BeautifulSoup"),
        (r"search_and_crawl\s*\(", "search_and_crawl()"),
        (r"\bsearch\s*\(\s*[\"']", "search()"),
        (r"\bcrawl\s*\(\s*[\"']", "crawl()"),
    ]
    for pattern, name in forbidden_in_researcher:
        if re.search(pattern, code):
            return (
                f"你的代码包含被禁止的 `{name}` 调用。"
                "researcher agent 禁止手写 HTTP/HTML 解析代码，也不应调用 search()/crawl()。"
                "搜索已由系统预执行，请直接使用提示词中提供的 _SEARCH_RESULTS 和 _CRAWLED_RESULTS。"
                "请删除相关调用代码后重新生成。"
            )
    if "_SEARCH_RESULTS" not in code:
        return (
            "你的代码未使用预获取的搜索数据 `_SEARCH_RESULTS`。"
            "请将提示词中提供的 _SEARCH_RESULTS 和 _CRAWLED_RESULTS 数据定义"
            "直接包含在你的代码中，然后用 DuckDB 存储并输出报告。"
        )
    return None


def _patch_reportlab_cjk_support(code: str) -> str:
    if "reportlab" not in code or ".pdf" not in code.lower():
        return code
    if CJK_FONT_HELPER_MARKER in code:
        return code

    patched = code
    if "from reportlab.pdfbase.cidfonts import UnicodeCIDFont" not in patched:
        if "from reportlab.pdfbase import pdfmetrics" in patched:
            patched = patched.replace(
                "from reportlab.pdfbase import pdfmetrics",
                "from reportlab.pdfbase import pdfmetrics\nfrom reportlab.pdfbase.cidfonts import UnicodeCIDFont",
                1,
            )
        else:
            patched = "from reportlab.pdfbase import pdfmetrics\nfrom reportlab.pdfbase.cidfonts import UnicodeCIDFont\n" + patched

    helper = """
# CJK_FONT_RUNTIME_PATCH_V1
def _ensure_cjk_font_runtime():
    import os

    candidates = [
        ("MicrosoftYaHei", r"C:\\Windows\\Fonts\\msyh.ttc"),
        ("MicrosoftYaHeiUI", r"C:\\Windows\\Fonts\\msyh.ttf"),
        ("SimHei", r"C:\\Windows\\Fonts\\simhei.ttf"),
        ("SimSun", r"C:\\Windows\\Fonts\\simsun.ttc"),
        ("NotoSansCJKsc", r"C:\\Windows\\Fonts\\NotoSansCJK-Regular.ttc"),
    ]
    for font_name, font_path in candidates:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                try:
                    pdfmetrics.registerFontFamily(
                        font_name,
                        normal=font_name,
                        bold=font_name,
                        italic=font_name,
                        boldItalic=font_name,
                    )
                except Exception:
                    pass
                return font_name
            except Exception:
                continue
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        try:
            pdfmetrics.registerFontFamily(
                "STSong-Light",
                normal="STSong-Light",
                bold="STSong-Light",
                italic="STSong-Light",
                boldItalic="STSong-Light",
            )
        except Exception:
            pass
        return "STSong-Light"
    except Exception:
        return ""

_CJK_FONT_NAME = _ensure_cjk_font_runtime()
""".strip()

    anchor = "from reportlab.pdfbase.ttfonts import TTFont"
    if anchor in patched:
        patched = patched.replace(anchor, anchor + "\n" + helper, 1)
    else:
        patched = helper + "\n\n" + patched

    direct_replacements = [
        ('return "Helvetica"', 'return _CJK_FONT_NAME or "Helvetica"'),
        ("return 'Helvetica'", "return _CJK_FONT_NAME or 'Helvetica'"),
        ('chinese_font_name = "Helvetica"', 'chinese_font_name = _CJK_FONT_NAME or "Helvetica"'),
        ("chinese_font_name = 'Helvetica'", "chinese_font_name = _CJK_FONT_NAME or 'Helvetica'"),
        ('FONT_NAME = "Helvetica"', 'FONT_NAME = _CJK_FONT_NAME or "Helvetica"'),
        ("FONT_NAME = 'Helvetica'", "FONT_NAME = _CJK_FONT_NAME or 'Helvetica'"),
        ("('FONTNAME', (0, 0), (-1, -1), 'Helvetica')", "('FONTNAME', (0, 0), (-1, -1), _CJK_FONT_NAME or 'Helvetica')"),
        ('("FONTNAME", (0, 0), (-1, -1), "Helvetica")', '("FONTNAME", (0, 0), (-1, -1), _CJK_FONT_NAME or "Helvetica")'),
    ]
    for before, after in direct_replacements:
        patched = patched.replace(before, after)

    patched = re.sub(
        r"fontName\s*=\s*['\"]Helvetica(?:-Bold)?['\"]",
        "fontName=_CJK_FONT_NAME or 'Helvetica'",
        patched,
    )
    patched = re.sub(
        r"setFont\(\s*['\"]Helvetica(?:-Bold)?['\"]\s*,",
        "setFont(_CJK_FONT_NAME or 'Helvetica',",
        patched,
    )
    patched = re.sub(
        r'pdfmetrics\.registerFont\s*\(\s*TTFont\s*\(\s*[\'"]STSong-Light[\'"]\s*,\s*[\'"]STSong-Light[\'"]\s*\)\s*\)',
        'pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))',
        patched,
    )
    return patched


def _execute_local_python_code(code: str, thread_id: str, attempt_index: int) -> dict:
    ensure_directories()
    run_dir = _make_run_dir(thread_id, attempt_index)
    code_path = run_dir / "generated_code.py"
    result_path = run_dir / "result.json"
    code_path.write_text(code, encoding="utf-8")
    _copy_duckdb_to_workdir(run_dir)
    if SEARCH_SERVICE_PATH.exists():
        shutil.copy2(SEARCH_SERVICE_PATH, run_dir / "search_service.py")

    try:
        completed = subprocess.run(
            [sys.executable, str(LOCAL_RUNNER_PATH), str(code_path), str(result_path)],
            cwd=str(run_dir),
            capture_output=True,
            text=True,
            timeout=LOCAL_EXEC_TIMEOUT_SECONDS,
            check=False,
            env={
                **os.environ,
                "PYTHONPATH": str(PROJECT_ROOT) + (
                    os.pathsep + os.environ["PYTHONPATH"] if os.environ.get("PYTHONPATH") else ""
                ),
            },
        )
        if result_path.exists():
            payload = json.loads(result_path.read_text(encoding="utf-8"))
        else:
            payload = {
                "ok": completed.returncode == 0,
                "stdout": completed.stdout,
                "stderr": completed.stderr,
                "traceback": "",
                "error_type": "",
                "generated_files": [],
            }
    except subprocess.TimeoutExpired:
        payload = {
            "ok": False,
            "stdout": "",
            "stderr": "",
            "traceback": f"TimeoutExpired: 本地执行超过 {LOCAL_EXEC_TIMEOUT_SECONDS} 秒后被终止。",
            "error_type": "TimeoutExpired",
            "generated_files": [],
        }

    _sync_duckdb_back(run_dir)
    artifact_info = _collect_local_artifacts(
        run_dir,
        payload.get("stdout", ""),
        payload.get("generated_files", []),
    )
    traceback_text = payload.get("traceback", "")
    stderr_text = payload.get("stderr", "")
    error_output = traceback_text or stderr_text

    return {
        "result_output": artifact_info["normalized_output"],
        "error_output": error_output,
        "chart_path": artifact_info["chart_path"],
        "chart_png_path": artifact_info["chart_png_path"],
        "cleaned_file_path": artifact_info["cleaned_file_path"],
        "report_path": artifact_info["report_path"],
        "pdf_report_path": artifact_info["pdf_report_path"],
        "pptx_report_path": artifact_info["pptx_report_path"],
        "validation_report_path": artifact_info["validation_report_path"],
        "artifacts": artifact_info["artifacts"],
        "stdout": payload.get("stdout", ""),
        "stderr": stderr_text,
        "traceback": traceback_text,
        "error_type": payload.get("error_type", ""),
        "execution_backend": "local_subprocess",
    }


# 哪些错误关键词判定为 sandbox 已死、应重建
_DEAD_SANDBOX_PATTERNS = (
    "port is not open", "502", "503", "504",
    "connection refused", "connection reset", "remote end closed",
    "sandbox not found", "sandbox is not running",
)


def _is_dead_e2b_error(exc: BaseException) -> bool:
    msg = (str(exc) or "").lower()
    return any(p.lower() in msg for p in _DEAD_SANDBOX_PATTERNS)


def _try_e2b_execution(code: str, e2b_api_key: str, duckdb_path: Path) -> dict:
    """带重试的 E2B 执行：撞 502/dead sandbox 自动重建并重跑，最多重试 2 次。"""
    template_id = os.environ.get("E2B_TEMPLATE_ID", "code-interpreter-v1")
    try:
        sandbox_timeout = int(os.environ.get("SPECTRA_E2B_TIMEOUT", "1800"))
    except (TypeError, ValueError):
        sandbox_timeout = 1800

    last_exc: Exception | None = None
    max_retry = 2
    for attempt in range(max_retry + 1):
        try:
            with Sandbox.create(
                api_key=e2b_api_key, template=template_id, timeout=sandbox_timeout,
            ) as sandbox:
                if duckdb_path.exists():
                    sandbox.files.write("data.duckdb", duckdb_path.read_bytes())
                if SEARCH_SERVICE_PATH.exists():
                    sandbox.files.write("search_service.py", SEARCH_SERVICE_PATH.read_bytes())

                execution = sandbox.run_code(code)
                stdout_text = execution.text or ""
                error_text = ""
                error_type = ""
                if execution.error:
                    error_type = execution.error.name
                    error_text = f"{execution.error.name}: {execution.error.value}\n{execution.error.traceback}"

                artifact_info = _collect_e2b_artifacts(sandbox, stdout_text)

                if duckdb_path.exists():
                    try:
                        duckdb_path.write_bytes(sandbox.files.read_bytes("data.duckdb"))
                    except Exception:
                        pass

                return {
                    "result_output": artifact_info["normalized_output"],
                    "error_output": error_text,
                    "chart_path": artifact_info["chart_path"],
                    "chart_png_path": artifact_info["chart_png_path"],
                    "cleaned_file_path": artifact_info["cleaned_file_path"],
                    "report_path": artifact_info["report_path"],
                    "pdf_report_path": artifact_info["pdf_report_path"],
                    "pptx_report_path": artifact_info["pptx_report_path"],
                    "validation_report_path": artifact_info["validation_report_path"],
                    "artifacts": artifact_info["artifacts"],
                    "stdout": stdout_text,
                    "stderr": "",
                    "traceback": error_text,
                    "error_type": error_type,
                    "execution_backend": "e2b",
                }
        except Exception as exc:
            last_exc = exc
            if _is_dead_e2b_error(exc) and attempt < max_retry:
                print(
                    f"[Executor] E2B sandbox 撞 502/dead，第 {attempt + 1} 次重建并重试: {exc}"
                )
                continue
            raise

    # 所有重试都失败
    raise last_exc if last_exc else RuntimeError("E2B retry budget exhausted")


def _execute_python_code(code: str, thread_id: str, attempt_index: int) -> dict:
    if LOCAL_ONLY_MARKER in code:
        return _execute_local_python_code(code, thread_id, attempt_index)
    e2b_api_key = os.environ.get("E2B_API_KEY", "").strip()
    if e2b_api_key:
        # 整体超时（沙盒 + 代码执行 + 产物拉取）：默认 600s。
        # 之前是 30s——对于带 LLM 思考时间或图表渲染的代码远不够，会让 E2B 路径几乎必然超时
        # fallback 到本地子进程，弄得"E2B 老是失败"。
        try:
            wrapper_timeout = int(os.environ.get("SPECTRA_E2B_RUN_TIMEOUT", "600"))
        except (TypeError, ValueError):
            wrapper_timeout = 600
        try:
            with ThreadPoolExecutor(max_workers=1) as pool:
                future = pool.submit(_try_e2b_execution, code, e2b_api_key, DUCKDB_PATH)
                return future.result(timeout=wrapper_timeout)
        except FutureTimeoutError:
            print(f"[Executor] E2B 沙盒执行超时 ({wrapper_timeout}s)，回退到本地子进程执行。")
        except Exception as exc:
            print(f"[Executor] E2B 沙盒执行失败，回退到本地子进程执行: {exc}")
    return _execute_local_python_code(code, thread_id, attempt_index)


def _strip_artifact_markers(text: str) -> str:
    cleaned_lines = []
    for line in text.splitlines():
        if line.startswith(("CHART_GENERATED:", "CHART_PNG_GENERATED:", "CLEANED_DATA_GENERATED:", "REPORT_GENERATED:")):
            continue
        cleaned_lines.append(line)
    return "\n".join(cleaned_lines).strip()


def _researcher_has_meaningful_results(state: dict, clean_output: str) -> bool:
    artifact_text = "\n".join(
        str(item.get("path", ""))
        for item in (state.get("artifacts") or [])
        if isinstance(item, dict)
    )
    combined = "\n".join([
        clean_output or "",
        state.get("execution_result", "") or "",
        state.get("last_stdout", "") or "",
        artifact_text,
    ])
    zero_result_markers = [
        "_SEARCH_RESULTS = []",
        "所有搜索源均无结果",
        "DuckDuckGo + Google 均无结果",
        "未获取到任何有效结果",
        "未获取到任何有效爬取内容",
        "未检测到明确表述",
        "未获取到有效信源",
        "共获取 0 个",
        "共获取 0 条",
        "获取 0 条候选",
        "搜索未返回结果",
    ]
    return not any(marker in combined for marker in zero_result_markers)


def _validate_execution(state: dict) -> dict:
    sender = state.get("sender", "analyzer")
    runtime_error = (state.get("last_traceback") or state.get("last_stderr") or "").strip()
    clean_output = _strip_artifact_markers(state.get("execution_result", "") or "")
    chart_path = state.get("chart_path")
    chart_png_path = state.get("chart_png_path")
    cleaned_file_path = state.get("cleaned_file_path")
    report_path = state.get("report_path")
    pdf_report_path = state.get("pdf_report_path")
    pptx_report_path = state.get("pptx_report_path")
    validation_report_path = state.get("validation_report_path")

    has_report = bool(pdf_report_path or report_path or pptx_report_path)
    issues: list[str] = []
    if runtime_error:
        issues.append(f"运行时报错:\n{runtime_error}")
    if sender == "cleaner" and not cleaned_file_path:
        issues.append("清洗任务未导出清洗后的 Excel 文件。")
    if sender == "cleaner" and cleaned_file_path:
        # 强校验：清洗产物必须有至少 1 行数据，否则视为清洗失败（典型场景：
        # LLM 误把 search_results 等历史表当成目标，导致 query 空、写出只有表头的 xlsx）
        try:
            from openpyxl import load_workbook
            cleaned_abs = (ARTIFACTS_DIR / cleaned_file_path).resolve()
            if cleaned_abs.exists():
                wb = load_workbook(cleaned_abs, read_only=True, data_only=True)
                ws = wb.active
                # max_row 包含表头行；扣掉后判断
                data_row_count = max(0, (ws.max_row or 0) - 1)
                wb.close()
                if data_row_count == 0:
                    issues.append(
                        "清洗后的 Excel 文件只有表头、零行数据。常见原因：写错了目标表名，"
                        "或 SQL/筛选条件导致结果为空。请重新检查 task 中提到的真实表名"
                        "（不要查询 `search_results` 等其它表）并修复 SQL。"
                    )
        except Exception as exc:
            # openpyxl 读不开就报错；正常 path 已经在 _collect_*_artifacts 阶段验证过文件可读
            issues.append(f"清洗后的 Excel 文件无法读取以校验行数：{exc}")
    if sender in {"reporter", "planner", "form_filler"} and not has_report:
        if chart_png_path or chart_path:
            issues.append("报告任务生成了图表文件，但未输出 PDF/DOCX 报告（图表未嵌入文档）。请使用 reportlab 生成 PDF 并用 Image 嵌入图表。")
        else:
            issues.append("报告任务未生成可下载的 PDF、DOCX 或 PPTX 报告。")

    # PPTX 专项校验：文件存在且能被 python-pptx 重新打开
    if pptx_report_path:
        pptx_abs = (ARTIFACTS_DIR / pptx_report_path).resolve()
        if not pptx_abs.exists():
            issues.append(f"PPTX 产物文件不存在: {pptx_report_path}")
        else:
            try:
                from pptx import Presentation
                Presentation(str(pptx_abs))
            except Exception as pptx_e:
                issues.append(f"PPTX 校验失败（无法用 python-pptx 打开）: {pptx_e}")
    if sender in {"reporter", "planner"} and clean_output:
        text_ref_patterns = [r"\[图表已生成:", r"CHART_GENERATED:", r"CHART_PNG_GENERATED:"]
        if any(re.search(p, clean_output) for p in text_ref_patterns):
            issues.append("报告仅以文字引用图表文件而未嵌入图片。请使用 reportlab.platypus.Image 将图表 PNG 嵌入 PDF 文档中，而不是仅打印文件路径标记。")
    if sender == "form_filler" and not validation_report_path:
        issues.append("PDF 自动填充任务未生成校验报告。")
    if sender in {"visualizer", "predictor"} and not (chart_path or chart_png_path):
        issues.append("可视化相关任务未生成图表产物。")
    if sender in {"analyzer", "predictor", "reporter", "planner", "form_filler", "researcher"} and not clean_output:
        issues.append("分析/预测任务没有输出可供用户阅读的结论。")
    if sender == "researcher" and not runtime_error and not _researcher_has_meaningful_results(state, clean_output):
        issues.append("联网搜索未获取到有效外部结果，当前输出仅包含空结果或无信源提示。")
    if not runtime_error and not clean_output and not state.get("artifacts"):
        issues.append("代码虽然运行结束，但没有产出任何结果或文件。")

    diagnostic_text = "\n\n".join(issues).strip()
    return {
        "ok": not issues,
        "diagnostic": diagnostic_text or "执行通过校验。",
        "error_category": _classify_error(state.get("last_error_type", ""), diagnostic_text),
        "clean_output": clean_output,
    }


def executor_node(state: dict) -> dict:
    code = state.get("generated_code", "") or ""
    code_type = state.get("code_type", "") or ""
    sender = state.get("sender", "analyzer") or "analyzer"
    retry_count = state.get("retry_count", 0)
    attempt_index = retry_count + 1
    thread_id = str(uuid.uuid4())
    configurable = state.get("configurable") if isinstance(state.get("configurable"), dict) else None
    if configurable and configurable.get("thread_id"):
        thread_id = str(configurable["thread_id"])

    if not code:
        return {
            "attempt_index": attempt_index,
            "execution_result": "",
            "last_stdout": "",
            "last_stderr": "",
            "last_traceback": "未生成任何代码。",
            "last_error_type": "MissingCode",
            "diagnostic": "未生成任何代码。",
            "execution_backend": "none",
            "artifacts": [],
        }

    if code_type != "python":
        return {
            "attempt_index": attempt_index,
            "execution_result": "",
            "last_stdout": "",
            "last_stderr": "",
            "last_traceback": f"不支持的代码类型: {code_type}",
            "last_error_type": "UnsupportedCodeType",
            "diagnostic": "不支持的代码类型。",
            "execution_backend": "none",
            "artifacts": [],
        }

    code = _patch_reportlab_cjk_support(code)
    preflight_error = _preflight_check(code)
    researcher_skill_active = (
        state.get("selected_skill_capability") == "web_search"
        or state.get("selected_skill_name") == "web-research"
    )
    if not preflight_error and ((sender or "").lower() == "researcher" or researcher_skill_active):
        preflight_error = _preflight_researcher_check(code)
    if preflight_error:
        return {
            "attempt_index": attempt_index,
            "execution_result": "",
            "last_stdout": "",
            "last_stderr": "",
            "last_traceback": preflight_error,
            "last_error_type": "PreflightBlocked",
            "diagnostic": preflight_error,
            "execution_backend": "preflight",
            "artifacts": [],
        }

    ensure_directories()
    print(f"[Executor] 开始执行 {sender} 生成的代码 (第 {attempt_index} 轮) ...")
    exec_result = _execute_python_code(code, thread_id, attempt_index)

    return {
        "attempt_index": attempt_index,
        "execution_result": exec_result["result_output"],
        "chart_path": exec_result["chart_path"],
        "chart_png_path": exec_result["chart_png_path"],
        "cleaned_file_path": exec_result["cleaned_file_path"],
        "report_path": exec_result["report_path"],
        "pdf_report_path": exec_result["pdf_report_path"],
        "pptx_report_path": exec_result["pptx_report_path"],
        "validation_report_path": exec_result["validation_report_path"],
        "artifacts": exec_result["artifacts"],
        "last_stdout": exec_result["stdout"],
        "last_stderr": exec_result["stderr"],
        "last_traceback": exec_result["traceback"],
        "last_error_type": exec_result["error_type"],
        "execution_backend": exec_result["execution_backend"],
        "selected_skill_name": state.get("selected_skill_name"),
        "selected_skill_path": state.get("selected_skill_path"),
        "selected_skill_description": state.get("selected_skill_description"),
        "selected_skill_capability": state.get("selected_skill_capability"),
        "skill_auto_created": state.get("skill_auto_created"),
        "execution_mode": state.get("execution_mode"),
        "fallback_source": state.get("fallback_source"),
    }


def validator_node(state: dict) -> dict:
    validation = _validate_execution(state)
    retry_count = state.get("retry_count", 0)
    max_retries = state.get("max_retries", MAX_GRAPH_RETRIES)
    clean_output = validation["clean_output"] or "代码执行完成。"

    if validation["ok"]:
        artifact_lines = []
        if state.get("chart_path"):
            artifact_lines.append(f"- 图表 HTML：`{state['chart_path']}`")
        if state.get("chart_png_path"):
            artifact_lines.append(f"- 图表 PNG：`{state['chart_png_path']}`")
        if state.get("cleaned_file_path"):
            artifact_lines.append(f"- 清洗文件：`{state['cleaned_file_path']}`")
        if state.get("report_path"):
            artifact_lines.append(f"- 报告：`{state['report_path']}`")
        if state.get("pdf_report_path"):
            artifact_lines.append(f"- PDF 报告：`{state['pdf_report_path']}`")
        if state.get("pptx_report_path"):
            artifact_lines.append(f"- PPTX 报告：`{state['pptx_report_path']}`")
        if state.get("validation_report_path"):
            artifact_lines.append(f"- 校验报告：`{state['validation_report_path']}`")

        reply = clean_output
        if artifact_lines:
            reply = f"{reply}\n\n已生成以下产物：\n" + "\n".join(artifact_lines)

        if clean_output.strip():
            try:
                from backend.memory import save_memory

                task_goal = _get_task_goal(state)
                if task_goal:
                    save_memory(task_goal, clean_output)
            except Exception as exc:
                print(f"记忆保存失败: {exc}")

        return {
            "validation_passed": True,
            "diagnostic": validation["diagnostic"],
            "reply": reply,
            "messages": [AIMessage(content=reply)],
        }

    diagnostic = validation["diagnostic"]
    if retry_count >= max_retries:
        final_reply = (
            f"系统已完成 {retry_count + 1} 轮执行，并尝试自动修复 {retry_count} 次，但仍未达到预期结果。\n\n"
            f"最后一次诊断如下：\n{diagnostic}"
        )
        return {
            "validation_passed": False,
            "diagnostic": diagnostic,
            "reply": final_reply,
            "messages": [AIMessage(content=final_reply)],
        }

    interim_reply = f"第 {retry_count + 1} 轮执行未通过校验，系统正在根据反馈自动修复。"
    return {
        "validation_passed": False,
        "next_node": "fixer",
        "diagnostic": diagnostic,
        "reply": interim_reply,
        "messages": [AIMessage(content=interim_reply)],
    }


def fixer_agent(
    code: str,
    error: str,
    schema: str = "",
    task_goal: str = "",
    stdout: str = "",
    stderr: str = "",
    traceback_text: str = "",
    attempt_index: int = 1,
    repair_history: list[str] | None = None,
) -> str:
    """LLM 根据诊断信息修复代码，返回修复后的完整 Python 代码（失败时返回空串）。"""
    llm = _create_llm()
    repair_history_text = "\n\n".join(repair_history or [])

    prompt = f"""你是一位资深 Python 调试专家，专精于 Pandas、Plotly、DuckDB 和数据分析领域。
你的任务是根据错误堆栈来诊断并修复一段执行失败的 Python 代码。

【原始业务目标】
{task_goal if task_goal else "（未提供）"}

【第 {attempt_index} 轮修复，历次诊断回顾（如有）】
{repair_history_text or "（无）"}

【数据库 schema（如有）】
{schema or "（未提供）"}

【失败的代码】
```python
{code}
```

【运行时错误 / 诊断】
{error}

【失败时的 stdout】
{stdout or "（无）"}

【失败时的 stderr】
{stderr or "（无）"}

【完整 traceback】
{traceback_text or "（无）"}

要求：
1. 直接给出修复后的完整可执行 Python 代码（不要省略），用 ```python ... ``` 包裹，代码块前可附 1-2 句修复说明。
2. 代码要保留原任务目标的 print 与 artifact 输出 marker（CHART_GENERATED / CHART_PNG_GENERATED / CLEANED_DATA_GENERATED / REPORT_GENERATED）。
3. 禁止 pip install / subprocess / os.system。环境已预装 pandas, numpy, duckdb, plotly, scikit-learn, statsmodels, openpyxl, kaleido, reportlab, python-docx 等。
4. 如果你无法修复，请返回**空字符串**，不要返回任何代码。
"""
    response = llm.invoke([SystemMessage(content=prompt), HumanMessage(content="请输出修复后的完整代码。")])
    content = getattr(response, "content", "") or ""

    # 提取代码块
    match = re.search(r"```(?:python)?\s*(.*?)```", content, re.DOTALL)
    if match:
        return match.group(1).strip()
    # 没找到代码块就直接返回空
    return ""


def fixer_node(state: dict) -> dict:
    """根据 validator 的诊断让 LLM 生成修复后的代码。"""
    retry_count = state.get("retry_count", 0)
    next_attempt = retry_count + 2
    current_code = state.get("generated_code", "") or ""
    diagnostic = state.get("diagnostic", "") or "代码执行失败，但未返回详细诊断。"
    task_goal = _get_task_goal(state)
    repair_history = list(state.get("repair_history", []) or [])

    # schema 由调用方按需注入，受 table_scope 约束
    try:
        from backend.request_context import get_table_scope
        from backend.db_utils import get_scoped_database_schema
        scope = get_table_scope()
        if scope is not None and len(scope) > 0:
            schema = get_scoped_database_schema(scope)
        else:
            schema = get_scoped_database_schema()  # 无 scope 时返回全库 schema
    except Exception:
        schema = ""

    fixed_code = fixer_agent(
        code=current_code,
        error=diagnostic,
        schema=schema,
        task_goal=task_goal,
        stdout=state.get("last_stdout", "") or "",
        stderr=state.get("last_stderr", "") or "",
        traceback_text=state.get("last_traceback", "") or "",
        attempt_index=next_attempt,
        repair_history=repair_history,
    )
    if not fixed_code.strip():
        fixed_code = current_code

    fix_summary = f"自动修复 Agent 已生成第 {next_attempt} 轮代码，准备重新执行。"
    repair_history.append(diagnostic[:1000])
    return {
        "generated_code": fixed_code,
        "code_type": "python",
        "retry_count": retry_count + 1,
        "fix_summary": fix_summary,
        "repair_history": repair_history,
        "messages": [AIMessage(content=fix_summary)],
    }
